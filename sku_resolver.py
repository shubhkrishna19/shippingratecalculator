from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from asset_paths import DIMENSIONS_WORKBOOK, SKU_ALIAS_WORKBOOK


@dataclass
class WeightMatch:
    matched: bool
    mtp_sku: Optional[str] = None
    mtp_name: Optional[str] = None
    weight_kg: Optional[float] = None
    weight_grams: Optional[float] = None
    matched_by: Optional[str] = None
    candidate_used: Optional[str] = None
    detail: Optional[str] = None


class DimensionsCatalog:
    def __init__(
        self,
        dimensions_path: Optional[Path] = None,
        alias_path: Optional[Path] = None,
        cleanup_suffixes: Optional[List[str]] = None,
    ) -> None:
        self.dimensions_path = dimensions_path or DIMENSIONS_WORKBOOK
        self.alias_path = alias_path or SKU_ALIAS_WORKBOOK
        self.cleanup_suffixes = cleanup_suffixes or []
        self.dimensions: Dict[str, Dict[str, Any]] = {}
        self.alias_maps: Dict[str, Dict[str, Dict[str, Optional[str]]]] = {}
        self.reload(cleanup_suffixes=self.cleanup_suffixes)

    def reload(self, cleanup_suffixes: Optional[List[str]] = None) -> None:
        if cleanup_suffixes is not None:
            self.cleanup_suffixes = cleanup_suffixes
        self.dimensions = self._load_dimensions()
        self.alias_maps = self._load_aliases()

    def resolve(self, candidates: List[Optional[str]]) -> WeightMatch:
        tried: List[str] = []

        for candidate in self._expand_candidates(candidates):
            if candidate in tried:
                continue
            tried.append(candidate)

            direct = self.dimensions.get(candidate)
            if direct:
                return WeightMatch(
                    matched=True,
                    mtp_sku=direct["mtp_sku"],
                    mtp_name=direct["mtp_name"],
                    weight_kg=direct["weight_kg"],
                    weight_grams=direct["weight_grams"],
                    matched_by="dimensions_master",
                    candidate_used=candidate,
                )

            for map_name, alias_map in self.alias_maps.items():
                alias_hit = alias_map.get(candidate)
                if not alias_hit:
                    continue

                mtp_sku = alias_hit["mtp_sku"]
                direct = self.dimensions.get(mtp_sku)
                if direct:
                    return WeightMatch(
                        matched=True,
                        mtp_sku=direct["mtp_sku"],
                        mtp_name=direct["mtp_name"] or alias_hit.get("mtp_name"),
                        weight_kg=direct["weight_kg"],
                        weight_grams=direct["weight_grams"],
                        matched_by=f"alias:{map_name}",
                        candidate_used=candidate,
                    )

                return WeightMatch(
                    matched=False,
                    mtp_sku=mtp_sku,
                    mtp_name=alias_hit.get("mtp_name"),
                    matched_by=f"alias:{map_name}",
                    candidate_used=candidate,
                    detail=f"Mapped to {mtp_sku}, but no dimensions row exists.",
                )

        detail = "No dimensions or alias match for candidates: " + ", ".join(tried[:12])
        return WeightMatch(matched=False, detail=detail)

    def _load_dimensions(self) -> Dict[str, Dict[str, Any]]:
        workbook = load_workbook(self.dimensions_path, read_only=True, data_only=True)
        sheet = workbook["Billing Dimensions"]
        dimensions: Dict[str, Dict[str, Any]] = {}

        for row in sheet.iter_rows(min_row=4, values_only=True):
            mtp_sku = self._clean_cell(row[0])
            if not mtp_sku:
                continue

            box_weights = [
                self._number(row[5]),
                self._number(row[9]),
                self._number(row[13]),
            ]
            total_weight_grams = sum(weight for weight in box_weights if weight is not None)

            if total_weight_grams <= 0:
                fallback_values = [self._number(value) for value in row[15:20]]
                total_weight_grams = max((value for value in fallback_values if value is not None), default=0)

            if total_weight_grams <= 0:
                continue

            dimensions[mtp_sku] = {
                "mtp_sku": mtp_sku,
                "mtp_name": None,
                "weight_grams": round(total_weight_grams, 3),
                "weight_kg": round(total_weight_grams / 1000, 3),
            }

        return dimensions

    def _load_aliases(self) -> Dict[str, Dict[str, Dict[str, Optional[str]]]]:
        if not self.alias_path.exists():
            return {}

        workbook = load_workbook(self.alias_path, read_only=True, data_only=True)
        sheet = workbook["Child SKUs - Alias Master"]

        maps = {
            "child_sku": {},
            "asin": {},
            "fk_fsn": {},
            "ul": {},
            "pf": {},
            "fnsku": {},
        }

        for row in sheet.iter_rows(min_row=2, values_only=True):
            mtp_sku = self._clean_cell(row[3])
            if not mtp_sku:
                continue

            mtp_name = self._clean_cell(row[4])
            payload = {"mtp_sku": mtp_sku, "mtp_name": mtp_name}

            values = {
                "child_sku": self._clean_cell(row[0]),
                "fnsku": self._clean_cell(row[9]),
                "asin": self._clean_cell(row[10]),
                "fk_fsn": self._clean_cell(row[11]),
                "ul": self._clean_cell(row[13]),
                "pf": self._clean_cell(row[14]),
            }

            for map_name, value in values.items():
                if value and value not in {"#N/A", "N/A"}:
                    maps[map_name][value] = payload

            if mtp_sku in self.dimensions and not self.dimensions[mtp_sku]["mtp_name"] and mtp_name:
                self.dimensions[mtp_sku]["mtp_name"] = mtp_name

        return maps

    def _expand_candidates(self, candidates: List[Optional[str]]) -> List[str]:
        expanded: List[str] = []

        for candidate in candidates:
            cleaned = self._clean_cell(candidate)
            if not cleaned:
                continue
            expanded.append(cleaned)
            expanded.extend(self._generate_suffix_variants(cleaned))

        return expanded

    def _generate_suffix_variants(self, value: str) -> List[str]:
        variants: List[str] = []
        queue = [value]
        seen = {value}

        while queue:
            current = queue.pop(0)
            for suffix in self.cleanup_suffixes:
                if not suffix:
                    continue
                if current.endswith(suffix):
                    next_value = current[: -len(suffix)].rstrip("-_")
                    if next_value and next_value not in seen:
                        seen.add(next_value)
                        queue.append(next_value)
                        variants.append(next_value)

        return variants

    @staticmethod
    def _clean_cell(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _number(value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

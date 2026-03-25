import csv
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from asset_paths import PINCODES_CSV, PRICE_WORKBOOK


def export_rate_calculator_from_workbook(
    workbook_path: Optional[Path] = None,
    target_path: Optional[Path] = None,
) -> Path:
    workbook_path = workbook_path or PRICE_WORKBOOK
    target_path = target_path or (workbook_path.parent / "rate_calculator.py")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Python Code"]

    lines: list[str] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        cell = row[0]
        if cell is None:
            continue
        lines.append(str(cell))

    content = "\n".join(lines).rstrip() + "\n"
    target_path.write_text(content, encoding="utf-8")
    return target_path


def export_pincodes_from_workbook(
    workbook_path: Optional[Path] = None,
    target_path: Optional[Path] = None,
) -> Path:
    workbook_path = workbook_path or PRICE_WORKBOOK
    target_path = target_path or PINCODES_CSV
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["M"]

    with target_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(list(row))

    return target_path
